# from ast import Continue
# , Global
# from msilib.schema import Error
# from posixpath import split
from openpyxl import Workbook
import requests
import json
from datetime import datetime
import time
import pandas as pd
import numpy as np
from numpy import append
# import turtle
import random
import xlsxwriter, os
from tabulate import tabulate
# from pathlib import Path


#import platform
# from cmath import nan
# from fcntl import F_GET_SEALS
# from importlib.resources import path
# import math
# import XlsxWriter
# import colorama

countRow = 2
ApiKeylist=[]

ApiKey ='0' 

headers = {
        "Accept": "application/json",
        "x-apikey": ApiKey
    }
urlVT = "https://www.virustotal.com/api/v3/{}/{}"
counterOfRequest = 0
ApiKeyIndex=1

userFileName='None'
# This methode change the API key that in APIKey List 
# It will return True if the key has been change other ways will return False
def changeAPIkey():
    global ApiKey
    global ApiKeyIndex
    global counterOfRequest
    if (ApiKeyIndex < len(ApiKeylist)):

        ApiKey = str(ApiKeylist[ApiKeyIndex])
        ApiKeyIndex +=1
        counterOfRequest=0

        return True
    else:
        return False


def getReputationVT(indicator, type):
#  This methode will return based on the bellow :
#  for 'file': 
#       errors : 
#           1- IF The API key has not reponed from Virus total side and cant be change as there is no more API Key
#               1- print("you have to change your API Key and save this API key/s to try after 24 h")
#               return {'found':'False'}
#           2- IF the number of requests (values from input) were more than allowed by the user premession 
#           print('Request rate limit exceeded. You are making more requests than allowed from VT')
#       IF there is no Errors : 
# #           Threat_Information = {
#             'found': 'False','type':'file', 'indicatorScor': None,
#             'threatlabel': None, 'indicatorName': 'N/A',
#             'typeTag': None, 'tags': [], 'importList': [],
#             'sha1': None, 'sha256': None, 'md5': None,
#             'typeExtension': None, 'sections': None, 'size': None
#         }

# for 'domain' and 'ip' and 'fileForTH': ->fileForTH = file for threat hunt 
#     errors: 
#         1- The number of requests are more than allowed by VT
#               print('Request rate limit exceeded. You are making more requests than allowed from VT')
#               print('Please, Change th API keys')
#               return Threat_Information
#         2- will return return Threat_Information[found]= :
#               - 'error403'
#               - 'error4'
#               for further reousns 
#       If there is no error will return for 'domain' input : 
#           reat_Information = {
#             'found': 'False', 'type':'domain','domain':indicator,
#             'indicatorScor': None,'creation_date':None,
#             'categories': None, 'tags': None,
#             'record_A': []
#                               }

#       If there is no error will return for 'ip' input :
#           Threat_Information = {
#           'found': 'False', 'type': type, 'IoC': indicator,
#           'indicatorScor': None, 'country': None, 'tags': None,
#           'sub-Net': None
#           }


#        If there is no error will return for 'fileForTH' input :
      
#           Threat_Information = {
#               'found': 'False', 'type': 'file', 'indicatorScor': None,'threatlabel':None, 'indicatorName': 'N/A',
#               'sha256': None, 'typeTag': None
#            }



    global counterOfRequest

    if counterOfRequest >= 240:
        if changeAPIkey():
            getReputationVT(indicator, type)
            pass
        else:
            print('Request rate limit exceeded. You are making more requests than allowed from VT\n')
            print('change your API Key and try again\n')
            return {'found':'False'}
    time.sleep(0.5)
    if type == 'file':
        
        Threat_Information = {
            'found': 'False','type':'file', 'indicatorScor': 'None',
            'threatlabel': 'None', 'indicatorName': 'None',
            'typeTag': 'None', 'tags': [], 'importList': [],
            'sha1': 'None', 'sha256': 'None', 'md5': 'None',
            'typeExtension': 'None', 'sections': 'None', 'size': 'None'
        }

        #Search in VT First for the MD5 reputation

        url = urlVT.format('files',indicator)
        # Get the resulte

        response = requests.get(url, headers=headers)
        counterOfRequest += 1


        # extrct it as json format


        if(response.status_code == 200):


                responseJson = json.loads(response.content)
               

                if (responseJson['data']['attributes']['last_analysis_stats']['malicious']>= 15):
                     Threat_Information['indicatorScor']= 'High-Malicious'
                elif (responseJson['data']['attributes']['last_analysis_stats']['malicious']>= 8):
                     Threat_Information['indicatorScor']= 'Malicious'
                elif (responseJson['data']['attributes']['last_analysis_stats']['malicious']>= 1):
                    if (responseJson['data']['attributes']['last_analysis_stats']['suspicious']>= 1):
                        Threat_Information['indicatorScor'] = 'Suspicious and May-Malicious'
                    else:
                        Threat_Information['indicatorScor']= 'May-Malicious'
                elif (responseJson['data']['attributes']['last_analysis_stats']['suspicious']>= 1):
                     Threat_Information['indicatorScor']= 'Suspiciouse'
                else:
                    Threat_Information['indicatorScor']= 'not-Detected'

                # if not Threat_Information['indicatorScor'] == 'not-Detected':
                #     
                # parse only information that needed
                Threat_Information['found'] = 'True'
                Threat_Information['sha1'] = responseJson['data']['attributes']['sha1']
                Threat_Information['sha256'] = responseJson['data']['attributes']['sha256']
                Threat_Information['md5'] = responseJson['data']['attributes']['md5']
                try:
                    Threat_Information['threatlabel'] = responseJson['data']['attributes']['popular_threat_classification']['suggested_threat_label']
                except:
                    pass

                try:
                    Threat_Information['indicatorName'] = responseJson['data']['attributes']['signature_info']['internal name']
                except:
                    pass
                if 'size'in responseJson['data']['attributes']:
                    Threat_Information['size'] = responseJson['data']['attributes']['size']
                try:
                    #try to check if the type_tag are available or not
                    Threat_Information['typeTag']= responseJson['data']['attributes']['type_tag']
                    Threat_Information['tags']= responseJson['data']['attributes']['tags']
                except:
                    pass

                try:
                    #try to check if the import list information are available or not
                    Threat_Information['importList']= responseJson['data']['attributes']['pe_info']['import_list']
                except:
                    pass

                try:
                    #try to check if the ext available or not
                    Threat_Information['typeExtension'] = responseJson['data']['attributes']['type_extension']
                except:
                    pass

        elif response.status_code == 204 or response.status_code == 429:
            if not changeAPIkey():
                print('Request rate limit exceeded. You are making more requests than allowed from VT\n')
                print('change your API Key and try again\n')
                return Threat_Information
            else:
                getReputationVT(indicator, type)

        elif response.status_code == '403':
            print('Forbidden. You dont have enough privileges to make the request.')
            return Threat_Information
        elif response.status_code == '404':
            print ('check your input type / not fount')
            return Threat_Information
        else:
            print('Unowkn error(check your API key, Number of request, network connection)')
            print( response.status_code)
            return Threat_Information

    # if the type search for domain reputation
    elif type == 'domain':

        Threat_Information = {
            'found': 'False', 'type':'domain','domain':indicator,
            'indicatorScor': 'None','creation_date':'None',
            'categories': 'None', 'tags': 'None',
            'record_A': []
                              }

        # Search in VT First for the IP reputation
        url = urlVT.format('domains',indicator)
        # Get the resulte

        response = requests.get(url, headers=headers)

        counterOfRequest += 1
        

        responseJson = json.loads(response.content)
        # print('Trace')
        # print(responseJson)
        if (response.status_code == 200):

            # Assign a Score for this indicator

            if (responseJson['data']['attributes']['last_analysis_stats']['malicious'] >= 15):
                Threat_Information['indicatorScor'] = 'High-Malicious'
            elif (responseJson['data']['attributes']['last_analysis_stats']['malicious'] >= 8):
                Threat_Information['indicatorScor'] = 'Malicious'
            elif (responseJson['data']['attributes']['last_analysis_stats']['malicious'] >= 1):
                if (responseJson['data']['attributes']['last_analysis_stats']['suspicious'] >= 1):
                    Threat_Information['indicatorScor'] = 'Suspicious and May-Malicious'
                else:
                    Threat_Information['indicatorScor'] = 'May-Malicious'
            elif (responseJson['data']['attributes']['last_analysis_stats']['suspicious'] >= 1):
                Threat_Information['indicatorScor'] = 'Suspiciouse'
            else:
                Threat_Information['indicatorScor'] = 'not-Detected'

            if not Threat_Information['indicatorScor'] == 'not-Detected':
                Threat_Information['found'] = 'True'
            try:
                ts = int(responseJson['data']['attributes']['creation_date'])
                Threat_Information['creation_date'] = datetime.utcfromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
            except:
                pass


            try:
                # try to check if the categories are available
                Threat_Information['categories'] = responseJson['data']['attributes']['categories']
            except:
                pass


            try:
                # try to check if the type_tag are available or not
                Threat_Information['tags'] = responseJson['data']['attributes']['tags']
            except:
                pass


            try:
                # try to check if the categories are available
                last_dns_records = responseJson['data']['attributes']['last_dns_records']

                for i in last_dns_records:
                    if i['type'] == 'A':
                        Threat_Information['record_A'].append(i['value'])

            except:
                pass
            return Threat_Information

        
            #check if there is a data available for this hash
        elif response.status_code == '204':
            if not changeAPIkey():
                print('Request rate limit exceeded. You are making more requests than allowed from VT')
                print('Please, Change th API keys')
                Threat_Information['found']= 'error204'
                return Threat_Information
            else:
                getReputationVT(indicator, type)

        elif response.status_code == '403':
            Threat_Information['found']= 'error403'
            print('error')
            return Threat_Information
        
        elif 'error' in responseJson:
             Threat_Information['found'] = 'error'
             print('error')
             return Threat_Information
        else:  
            Threat_Information['found'] = 'error'
            print('error')
            return Threat_Information
    elif type == 'ip':

        Threat_Information = {
            'found': 'False', 'type': type, 'IoC': indicator,
            'indicatorScor': 'None', 'country': 'None', 'tags': 'None',
            'sub-Net': 'None'
        }
        # Search in VT First for the IP reputation
        url = urlVT.format('ip_addresses', indicator)
        try:
            response = requests.get(url, headers=headers)
        except:
             print('Please check your connection')
             return Threat_Information
        counterOfRequest += 1

        if (response.status_code == 200):
            response = json.loads(response.content)
            # Assign a Score for this indicator

            if (response['data']['attributes']['last_analysis_stats']['malicious'] >= 15):
                Threat_Information['indicatorScor'] = 'High-Malicious'
            elif (response['data']['attributes']['last_analysis_stats']['malicious'] >= 8):
                Threat_Information['indicatorScor'] = 'Malicious'
            elif (response['data']['attributes']['last_analysis_stats']['malicious'] >= 1):
                if (response['data']['attributes']['last_analysis_stats']['suspicious'] >= 1):
                    Threat_Information['indicatorScor'] = 'Suspicious and May-Malicious'
                else:
                    Threat_Information['indicatorScor'] = 'May-Malicious'
            elif (response['data']['attributes']['last_analysis_stats']['suspicious'] >= 1):
                Threat_Information['indicatorScor'] = 'Suspiciouse'
            else:
                Threat_Information['indicatorScor'] = 'not-Detected'

            # if not Threat_Information['indicatorScor'] == 'not-Detected':
            Threat_Information['found'] = 'True'
            # parse only information that needed
            if 'country' in response['data']['attributes']:
                Threat_Information ['country'] = response['data']['attributes']['country']

            if 'tags' in response['data']['attributes']:
                Threat_Information['tags'] = response['data']['attributes']['tags']

            if 'network' in response['data']['attributes']:
                Threat_Information['sub-Net'] = response['data']['attributes']['network']

            return Threat_Information
    #     CHECK THE reputation of an IP
        elif response.status_code == '204':
            if not changeAPIkey():
                print('Request rate limit exceeded. You are making more requests than allowed from VT\n')
                print('Please, Change th API keys\n')
                Threat_Information['found'] = 'error204'
                return Threat_Information
            else:
                getReputationVT(indicator, type)

        elif response.status_code == '403':
            Threat_Information['found']= 'error403'
            return Threat_Information
        else:
            Threat_Information['found'] = 'error'
            return Threat_Information

    # file for short info about files
    elif type == 'fileForTH':

            # Indicators_Threat_Information=[]
            # count= 0
            Threat_Information = {
                'found': 'False', 'type': 'file', 'indicatorScor': 'None','threatlabel':'None', 'indicatorName': 'N/A',
                'sha256': 'None', 'typeTag': 'None'
            }

            # Search in VT First for the MD5 reputation
            url = urlVT.format('files', indicator)

            try:
                response = requests.get(url, headers=headers)
            except:
                print('Please check your connection')
            # Get the resulte


            counterOfRequest += 1

            # extrct it as json format

            if (response.status_code == 200):

                responseJson = json.loads(response.content)
                #  For enhancing
                # Assign a Score for this indicator

                if (responseJson['data']['attributes']['last_analysis_stats']['malicious'] >= 15):
                    Threat_Information['indicatorScor'] = 'High-Malicious'
                elif (responseJson['data']['attributes']['last_analysis_stats']['malicious'] >= 8):
                    Threat_Information['indicatorScor'] = 'Malicious'
                elif (responseJson['data']['attributes']['last_analysis_stats']['malicious'] >= 1):
                    if (responseJson['data']['attributes']['last_analysis_stats']['suspicious'] >= 1):
                        Threat_Information['indicatorScor'] = 'Suspicious and May-Malicious'
                    else:
                        Threat_Information['indicatorScor'] = 'May-Malicious'
                elif (responseJson['data']['attributes']['last_analysis_stats']['suspicious'] >= 1):
                    Threat_Information['indicatorScor'] = 'Suspiciouse'
                else:
                    Threat_Information['indicatorScor'] = 'not-Detected'

                if not Threat_Information['indicatorScor'] == 'not-Detected':
                    Threat_Information['found'] = 'True'
                    # parse only information that needed
                    # Threat_Information['sha1'] = responseJson['data']['attributes']['sha1']
                    Threat_Information['sha256'] = responseJson['data']['attributes']['sha256']
                    # Threat_Information['md5'] = responseJson['data']['attributes']['md5']
                    try:
                        Threat_Information['threatlabel'] = responseJson['data']['attributes']['popular_threat_classification']['suggested_threat_label']
                    except:
                        pass
                    try:
                        Threat_Information['indicatorName'] = responseJson['data']['attributes']['signature_info']['internal name']
                    except:
                        pass
                    try:
                        Threat_Information['typeTag']= responseJson['data']['attributes']['type_tag']
                    except: 
                        pass 
            elif response.status_code == '204':
                if not changeAPIkey():
                    print('Request rate limit exceeded. You are making more requests than allowed from VT\n')
                    print('Please, Change th API keys\n')
                    Threat_Information['found'] = 'error204'
                    return Threat_Information
                else:
                    getReputationVT(indicator, type)

            elif response.status_code == '403':
                Threat_Information['found']= 'error403'
                return Threat_Information
            else:
                Threat_Information['found'] = 'error'
                return Threat_Information

    return Threat_Information


hashes = [

'd1bdc5aaa294b4c52678c4c60f052569',
'087951566fb77fe74909d4e4828dd4cb',
'8aacf26df235661245e98cb60e820f51',
'be0d32bb3a12896ff16e3f667eb4b644',
'f388391ca443056fd3b4cc733c3b61cd']


def getReputationList(indicators, type, checkertype):
# This methode get an input a list of IoCs (hashes, Ips, Domains) and type(wich IoCs in the list (domain, IP, hash) check type (reputation, TI)
# will return an array of dictionaryof the type you searched for it

    Indicators_Threat_Information=[]
    if (checkertype == 'reputation'):

        if isinstance(indicators, list):

            for i in indicators:
                result = getReputationVT(i,type)
                if (result['found'] == 'True'):
                   Indicators_Threat_Information.append(result)
                
    elif (checkertype == 'TI'):

        if isinstance(indicators, list):
            for i in indicators:
                resultVT= getReputationVT(i, type)
                result = getfileTIVT(i, type)

                # result['found']
                if 'found' in resultVT:
                    if resultVT['found']=='True':
                        Indicators_Threat_Information.append(resultVT)
                if result is list:
                    if len(result) >=1:
                        Indicators_Threat_Information.append(result)

def getfileTIVT(IoC, type, deepSearch):

# This methode to get collect all the Objected or digrams used in virustotal by the users and include the IoC that searched for
# for the file /domain /IPs will return a list of (Files, IPs, Domains) if has been sean and be malicious or suspicious 
# 
# for errors: 
#           will print one of the bellow errors and return an eempty array :
#            print("you have to change your API Key and save this API key/s to try after 24 h")
#            print("Not Exist")
#            else with threspons error 
# 
# for the objects, we have to kind of object: 
#       1- has acheck attrpuite wich leade to be maliciouse or not, if the object has a check we will name it as checker
# 
# In this stage we have two attribute ("checker :available or not" "deepsearch: True or false ") here we have four status, lets check it to see
# each output. 
# 
# checker 'available': 
#               - Deep Search 'True' -> Search for the finding "that found in Oject" (file,domain,hashe) in Viruse Total and get full response 
#               - if it not found or Deep Search 'False' , return the bellow 
#                       - file : 
#                           temp = {
#                                                     'found': 'True', 'type': 'file', 'indicatorScor': 'Malicious/Suspicious',
#                                                     'threatlabel': None, 'indicatorName': 'N/A',
#                                                     'typeTag': None, 'sha256': None,
#                                    }
#                
#                      - Domain : 
#                           temp = {
#                                            'found': 'True', 'type': 'domnain', 'indicatorScor': 'Malicious/Suspicious',
#                                            'IoC': nodes[i]['entity_id'],
#                                            'typeTag': None, 'relation': 'has been seen'
#                                   }
#                      
#                       - IP: 
#                          temp = {
#                                             'found': 'True', 'type': 'IP', 'indicatorScor': 'Malicious/Suspicious',
#                                             'IoC': nodes[i]['entity_id'],
#                                             'typeTag': None, 'relation': 'has been seen'
#                                   }
# cheker 'Not available': 
#                 - check from VT and for each object dont check more than 30 if it's not deep search . else there is no limit 
# 
#                       ALL resukte from VT based on the catogery -> temp 
#                  
# Now return array with disctionra as data set for all findings that found

    global counterOfRequest


    if type == 'file':
        checkTyp= 'files'
    elif type == 'domain':
        checkTyp= 'domains'
    elif type == 'ip':
        checkTyp='ip_addresses'
    else:
        return -1

    # print('search for :'+ IoC)

    url = "https://www.virustotal.com/api/v3/{}/{}/graphs?limit=10".format(checkTyp, IoC)

    relatedIoCs = []

    if (True):

        time.sleep(0.5)
        try:
            response = requests.get(url, headers=headers)
        except:
            print('check your connection')
            return -2

        counterOfRequest += 1

        responseCode= response.status_code
        response = response.text
            #check if there is a data available for this hash

        if responseCode != 200:
            if responseCode == 429 or responseCode == 204:
                    if changeAPIkey():
                        return getfileTIVT(IoC, type)

                    else:
                        print('Request rate limit exceeded. You are making more requests than allowed from VT\n')
                        print('change your API Key and try again\n')
                        return relatedIoCs

            elif responseCode == 404:
                # print("Not Exist\n")
                return relatedIoCs

            else:
                # print(response)
                return relatedIoCs



        response = json.loads(response)

        if (response ['meta']['count'] >=1 ):
            nodes = response['data'][0]['attributes']['nodes']
            nodesLen = len(nodes)
            # print('found relations ')
            # check if the Reputation  avilable for each file in nodes or not
            checker = False
            if "entity_attributes" in nodes[0]:

                checker =True

            if checker:
                # print("checker available")

                for i in range(1, nodesLen):
                        if 'entity_attributes' in nodes[i]:
                            if 'has_detections' in nodes [i] ['entity_attributes']:
                                if nodes [i] ['entity_attributes']['has_detections']:
                                    if (nodes [i]['type'] == 'file' ):
                                        # check from other resources befor check the virus total
                                        if deepSearch:
                                            temp2 = getReputationVT(nodes[i]['entity_id'], 'file')
                                            
                                            if temp2['found'] == 'True':
                                                relatedIoCs.append(temp2)
                                            else:
                                                temp = {
                                                    'found': 'True', 'type': 'file', 'indicatorScor': 'Malicious/Suspicious',
                                                    'threatlabel': 'None', 'indicatorName': 'N/A',
                                                    'typeTag': 'None', 'sha256': 'None',
                                                }
                                                # check if tags available
                                                if 'type_tag' in nodes[i]['entity_attributes']:
                                                    temp['typeTag'] = nodes[i]['entity_attributes']['type_tag']
                                                   
                                                temp['sha256'] = nodes[i]['entity_id']
                                               
                                                relatedIoCs.append(temp)
                                        else:
                                            
                                                temp = {
                                                    'found': 'True', 'type': 'file', 'indicatorScor': 'Malicious/Suspicious',
                                                     'threatlabel': 'N/A', 'indicatorName': 'N/A',
                                                     'typeTag': 'None', 'sha256': 'None',
                                                }
                                                    # check if tags available
                                                if 'type_tag' in nodes[i]['entity_attributes']:
                                                    temp['typeTag'] = nodes[i]['entity_attributes']['type_tag']
                                                  
                                                temp['sha256'] = nodes[i]['entity_id']
                                                
                                                relatedIoCs.append(temp)


                                    elif nodes[i]['type'] == 'domain':
                                        temp = {
                                            'found': 'True', 'type': 'domain', 'indicatorScor': 'Malicious/Suspicious',
                                            'IoC': nodes[i]['entity_id'],
                                            'typeTag': 'None', 'relation': 'has been seen'
                                        }
                                        
                                        if deepSearch:
                                            temp2 = getReputationVT(nodes[i]['entity_id'], 'domain')
                                            
                                            # print(temp2)
                                            if temp2['found'] == 'True':
                                               
                                                relatedIoCs.append(temp2)
                                            else:
                                                if 'type_tag' in nodes[i]['entity_attributes']:
                                                    temp['typeTag'] = nodes[i]['entity_attributes']['type_tag']
                                               
                                                relatedIoCs.append(temp)

                                        else:

                                            if 'type_tag' in nodes[i]['entity_attributes']:
                                                temp['typeTag'] = nodes[i]['entity_attributes']['type_tag']
                                            relatedIoCs.append(temp)


                                        # ip_address
                                    elif nodes[i]['type'] == 'ip_address' :
                                        temp = {
                                            'found': 'True', 'type': 'IP', 'indicatorScor': 'Malicious/Suspicious',
                                            'IoC': nodes[i]['entity_id'],
                                            'typeTag': 'None', 'relation': 'has been seen'
                                            }
                                        # should be search for further information
                                        if deepSearch:
                                            temp2 = getReputationVT(nodes[i]['entity_id'], 'ip')
                                            if temp['found'] == 'True':
                                                relatedIoCs.append(temp2)
                                            else : 
                                                if 'type_tag' in nodes[i]['entity_attributes']:
                                                    temp['typeTag'] = nodes[i]['entity_attributes']['type_tag']
                                            
                                                relatedIoCs.append(temp)

                                        else: 
                                           
                                            if 'type_tag' in nodes[i]['entity_attributes']:
                                                temp['typeTag'] = nodes[i]['entity_attributes']['type_tag']
                                            
                                            relatedIoCs.append(temp)
            else:
                counter=0
                for i in range(1, nodesLen):
                        if (nodes [i]['type'] == 'file'):
                            # check from other resources befor check the virus total
                            
                            if doseFileInMHR(nodes[i]['entity_id']):
                                temp = {
                                    'found': 'True', 'type': 'file', 'indicatorScor': 'None',
                                    'threatlabel': 'None', 'indicatorName': 'N/A',
                                    'typeTag': 'None', 'sha256': 'None',
                                }
                                #  t
                            #     We have to build it ASAP
                            else:
                                

                                if deepSearch:
                                    reputation = getReputationVT(nodes[i]['entity_id'], 'file')
                                    if reputation == 'None':
                                        return relatedIoCs

                                    if reputation['found'] == 'True' and not (reputation['indicatorScor'] == 'not-Detected'):
                                        
                                        relatedIoCs.append(reputation)

                                else:
                                    if counter >=30:
                                        break
                                    counter +=1
                                    reputation = getReputationVT(nodes[i]['entity_id'],'file')
                                    if reputation == 'None':
                                        return relatedIoCs

                                    if reputation['found'] == 'True' and not (reputation['indicatorScor'] == 'not-Detected' ):
                                        relatedIoCs.append(reputation)
                        elif nodes[i]['type'] == 'domain':
                            if not deepSearch: 
                                if counter >=30:
                                    break
                                counter +=1
                            domainResponse= getReputationVT(nodes[i]['entity_id'], 'domain')

                            # if reputation == None:
                            #     return relatedIoCs
                            # 'not-Detected'
                            if domainResponse['found'] == 'True' and not (reputation['indicatorScor'] == 'not-Detected' ) :
                                relatedIoCs.append(domainResponse)
                            elif domainResponse['found'] == 'error204':
                                print('Invalid API Key\n')
                        elif nodes[i]['type'] == 'ip_address':
                            if not deepSearch: 
                                if counter >=30:
                                    break
                                counter +=1
                            domainResponse = getReputationVT(nodes[i]['entity_id'], 'ip')
                            if 'error' in domainResponse:
                                    # print('error 2'+ domainResponse)
                                    continue
                            # if reputation == None:
                            #     return relatedIoCs
                            if domainResponse['found'] == 'True' and not (reputation['indicatorScor'] == 'not-Detected' ):
                                relatedIoCs.append(domainResponse)
                            elif domainResponse['found'] == 'error204':
                                print('Invalid API Key, change your API Key and try again\n')
                                

        else:
            pass

    return relatedIoCs

def doseFileInMHR(IoC):

    return False 

def readFile(path):
    global userFileName 
# This method to read type of files
# Excell Sheet and csv sheet only supported 
# if Error 1, the file not supported yet
    
    

    pathname, extension = os.path.splitext(path)
    filename = pathname.split('/')
    userFileName = filename
    # print(filename[-1])

    if not path.lower().endswith(('csv','xlsx','csv\'','xlsx\'', 'csv\"','xlsx\"')):
        return {'error':1}
    else: 
       
        # in IoCs will be a dictoneray 
        # keys {Hashes, Domains, IPs}
        # if the key not exist, that mean the user not upload the IoC we looking for
        IoCs = {'error':0}
        # try: 
        if path.lower().endswith(('csv')):
                file = pd.read_csv(path)
        else: 
                file = pd.read_excel(path)
        # except: 
            # print('The file not found')
        # reade and extract hashes, domains, ips from the File
        if 'Hashes' in file:
                #Add all hashes as one array into IoC dic
            hashesList = np.ravel(file['Hashes']).tolist()
            # Remove all null (nan) values frome the list
            hashesList = [item for item in hashesList if not(pd.isnull(item)) == True]
            IoCs['hashes']= hashesList
            
        # Parse all Domains in the Excell Sheet 
        if 'Domains' in file : 
            #Add all Domains as one array into IoC dic
             domainsList = np.ravel(file['Domains']).tolist()

            # Remove all null (nan) values frome the list
             domainsList = [item for item in domainsList if not(pd.isnull(item)) == True]
             IoCs['domains']= domainsList
               
        # Pares all IPs in the Excell Sheet
        if 'Ips' in file : 
            #Add all IPs as one array into IoC dic
            IPsList = np.ravel(file['Ips']).tolist()

            # Remove all null (nan) values frome the list
            IPsList = [item for item in IPsList if not(pd.isnull(item)) == True]
            IoCs['Ips']= IPsList

        # Excell Sheet
        

        return IoCs 

def readAPIKeyFile():
        global ApiKeylist
        global ApiKey
        global headers


        if not os.path.exists('api.xlsx'):
            return False
        else: 
            file = pd.read_excel('api.xlsx')
            if 'key' in file:
                APIList = np.ravel(file['key']).tolist()
                # print(APIList)
                APIList = [item for item in APIList if not(pd.isnull(item)) == True]
                if len(APIList)== 0:
                    return False
                # print(APIList)
                ApiKeylist.clear
                ApiKey = str(APIList[0])
                headers = {
                "Accept": "application/json",
                "x-apikey": ApiKey
                }
                ApiKeylist = APIList
                return True
        return False
            
def setAPIKeyFile(keys):
    if os.path.exists("api.xlsx"):
        os.remove("api.xlsx")
    try:
        countrows=2
        workbook = xlsxwriter.Workbook('api.xlsx')
        worksheet= workbook.add_worksheet('SHEHANH tool')
        worksheet.write('A1','key')

        for i in keys:
            if len(i)<9:
                continue
            worksheet.write('A'+str(countrows),str(i))
            countrows +=1

    finally: 
        workbook.close ()

def check(typeOfCheck, IoCs, Live):
    # test
    # if Live: 

    
    IoCsRP = []

    if typeOfCheck == 'r':
        if len(IoCs['hashes']) >= 1:
        # if '' in IoCs: 
            for i in IoCs['hashes']:
                if len(i) <=3:
                    continue
                temp = getReputationVT(i,'file')
                if temp['found'] == 'True':
                    if Live: 
                        IoCsRP.append({'input': i,'type': 'File', 'scor':temp['indicatorScor'],'sha256':temp['sha256']})
                    else: 
                        IoCsRP.append({'type':'hashe', 'score':temp['indicatorScor'], 
                                    'IoC':i, 'threatlabel': temp['threatlabel'], 'md5':temp['md5'], 'sha256': temp['sha256']})
                else : 
                    if Live: 
                        IoCsRP.append({'input': i,'type': 'File', 'scor':temp['indicatorScor'],'sha256':temp['sha256']})
                    else: 
                        IoCsRP.append({'type':'hashe', 'score':temp['indicatorScor'], 
                                    'IoC':i, 'threatlabel': temp['threatlabel'], 'md5':temp['md5'], 'sha256': temp['sha256']})
                        
        if len(IoCs['domains']) >= 1:
            for i in IoCs['domains']:
                if len(i) <=3:
                    continue
                temp = getReputationVT(i,'domain')
                # print(temp)
                if temp['found'] == 'True':
                    if Live: 
                        IoCsRP.append({'input': i,'type': 'Domain', 'scor':temp['indicatorScor'],'sha256':temp['record_A']})
                    else: 
                        IoCsRP.append({'type':'domain', 'score':temp['indicatorScor'], 
                                    'IoC':i, 'threatlabel': temp['record_A'], 'md5':'', 'sha256': '' })
                else:
                    if Live: 
                        IoCsRP.append({'input': i,'type': 'Domain', 'scor':temp['indicatorScor'],'sha256':temp['record_A']})
                    else: 
                        IoCsRP.append({'type':'domain', 'score':temp['indicatorScor'], 
                                    'IoC':i, 'threatlabel': temp['record_A'], 'md5':'', 'sha256': '' })
        if len(IoCs['Ips']) >= 1: 
            for i in IoCs['Ips']:
                if len(i) <=3:
                    continue
                temp = getReputationVT(i,'ip')
                # print(temp)
                if temp['found'] == 'True':
                    if Live: 
                        IoCsRP.append({'input': i,'type': 'IP', 'scor':temp['indicatorScor'],'sha256':temp['sub-Net']})
                    else: 
                        IoCsRP.append({'type':'IP', 'score':temp['indicatorScor'], 
                                    'IoC':i, 'threatlabel': str('country: '+ str(temp['country'])), 'md5':'', 'sha256': '' })
                else : 
                    if Live: 
                        IoCsRP.append({'input': i,'type': 'IP', 'scor':temp['indicatorScor'],'sha256':temp['sub-Net']})
                    else: 
                        IoCsRP.append({'type':'IP', 'score':temp['indicatorScor'], 
                                    'IoC':i, 'threatlabel': 'country: '+temp['country'], 'md5':'', 'sha256': '' })
        if Live : 
            displayLive(IoCsRP, 'r')
        else: 
            writeExcellFile(IoCsRP,'r')
                # writeExcellFile(IoCsRP,'r')
                
                # write it in File 

    # "Threat Huntting"

    elif typeOfCheck == 'th': 
        if len(IoCs['hashes']) >= 1: 
            for i in IoCs['hashes']:
                if len(i) <=3:
                    continue
                temp = getReputationVT(i,'file')
                if temp['found'] == 'True':
                    tempTI = getfileTIVT(i,'file',False)
                    
                    if Live: 
                        
                        IoCsRP.append({'IoC':i, 'type':'File', 'score':temp['indicatorScor'],'sha256': temp['sha256'], 'related': tempTI})
                                    #  'threatlabel': temp['threatlabel']
                    else: 
                   
                        IoCsRP.append({'IoC':i, 'type':'File', 'score':temp['indicatorScor'],
                                     'IoCName':temp['indicatorName'], 'threatlabel': temp['threatlabel'],
                                     'md5':temp['md5'], 'sha256': temp['sha256'],'typeExtension':temp['typeExtension'],
                                     'importList':temp['importList'],'size':temp['size'] ,
                                      'related': tempTI})
                else: 
                    if Live: 
                        
                        IoCsRP.append({'IoC':i, 'type':'File', 'score':'Not Found','sha256': '', 'related': ''})
                        # ,
                        #              'threatlabel': ''
                    else: 
                    
                        IoCsRP.append({'IoC':i, 'type':'File', 'score':'Not Found',
                                     'IoCName':'', 'threatlabel': '',
                                     'md5':'', 'sha256':'','typeExtension':'',
                                     'importList':'','size':'' ,
                                      'related': []})
        if len(IoCs['domains']) >= 1:
            for i in IoCs['domains']:
                if len(i) <=3: 
                    continue
                temp = getReputationVT(i,'domain')

                if temp['found'] == 'True':
                    tempTI = getfileTIVT(i,'domain',False)

                    if Live: 
                        IoCsRP.append({'IoC':i, 'type':'domain', 'score':temp['indicatorScor'],'sha256': temp['record_A'], 'related': tempTI})
                    else: 
                        # temp['related'] = tempTI
                        IoCsRP.append({'IoC':i, 'type':'domain', 'score':temp['indicatorScor'],
                                     'IoCName':'N/A', 'threatlabel': temp['categories'],
                                     'md5':temp['tags'], 'sha256':temp['record_A'], 'related': tempTI})
                else: 
                    if Live: 
                        IoCsRP.append({'IoC':i, 'type':'domain', 'score':'Not Found','sha256': i, 'related': ''})
                    else: 
                        # temp['related'] = tempTI
                        IoCsRP.append({'IoC':i, 'type':'domain', 'score':'Not Found',
                                     'IoCName':'', 'threatlabel':'',
                                     'md5':'', 'sha256':'', 'related': []})
        if len(IoCs['Ips']) >= 1:  

            for i in IoCs['Ips']:
                if len(i) <=3:
                    continue
                temp = getReputationVT(i,'ip')
                if temp['found'] == 'True':
                    tempTI = getfileTIVT(i,'ip',False)
                   
                    if Live: 
                        
                        IoCsRP.append({'IoC':i, 'type':'IP', 'score':temp['indicatorScor'],'sha256': temp['sub-Net'], 'related': tempTI})
                    else: 
                        
                        IoCsRP.append({'IoC':i, 'type':'IP', 'score':temp['indicatorScor'],
                                     'IoCName':'N/A', 'threatlabel': 'N/A',
                                     'md5':temp['tags'], 'sha256': temp['sub-Net'], 'related': tempTI})
                else: 
                    if Live: 
                        
                        IoCsRP.append({'IoC':i, 'type':'IP', 'score':'Not Found','sha256': '', 'related': ''})
                    else: 
                        IoCsRP.append({'IoC':i, 'type':'IP', 'score':'Not Found',
                                     'IoCName':'', 'threatlabel': '',
                                     'md5':'', 'sha256': '', 'related': ''})
        if Live: 
            displayLive(IoCsRP, 'th')
        else:
            writeExcellFile(IoCsRP, 'th')
    elif typeOfCheck == 'ti':
        if 'hashes' in IoCs: 
            for i in IoCs['hashes']:
                if len(i) <=3:
                    continue
                temp = getReputationVT(i,'file')
                if temp['found'] == 'True':
                    tempTI = getfileTIVT(i,'file',True)

                    IoCsRP.append({'IoC':i, 'type':'File', 'score':temp['indicatorScor'],
                            'IoCName':temp['indicatorName'],'typeTag':temp['typeTag'], 'tags':temp['tags'], 'threatlabel': temp['threatlabel'],
                            'md5':temp['md5'], 'sha256': temp['sha256'],'sha1':temp['sha1'],'typeExtension':temp['typeExtension'],
                            'importList':temp['importList'],'size':temp['size'],'sections': temp['sections'],
                            'related': tempTI})
                else: 
                    IoCsRP.append({'IoC':i, 'type':'File', 'score':'Not Found',
                            'IoCName':'','typeTag':'', 'tags':'', 'threatlabel': '',
                            'md5':'', 'sha256': '','sha1':'','typeExtension':'',
                            'importList':'','size':'','sections': '',
                            'related': []})
        if 'domains' in IoCs:
            for i in IoCs['domains']:
                if len(i) <=3: 
                    continue
                temp = getReputationVT(i,'domain')
                if temp['found'] == 'True':
                    tempTI = getfileTIVT(i,'domain',True)
                     
                    IoCsRP.append({'IoC':i, 'type':'domain', 'score':temp['indicatorScor'],
                                'IoCName':'', 'threatlabel': temp['categories'],'creation_date':temp['creation_date'],
                                'md5':temp['tags'], 'sha256':temp['record_A'], 'related': tempTI})
                else: 
                    IoCsRP.append({'IoC':i, 'type':'domain', 'score':'Not Found',
                                'IoCName':'', 'threatlabel':'','creation_date':'',
                                'md5':'', 'sha256':'', 'related': []})
        if 'Ips' in IoCs: 

            for i in IoCs['Ips']:
                if len(i) <=3:
                    continue
                temp = getReputationVT(i,'ip')
                if temp['found'] == 'True':
                    tempTI = getfileTIVT(i,'ip',True)
                    IoCsRP.append({'IoC':i, 'type':'IP', 'score':temp['indicatorScor'],
                                     'IoCName':'', 'threatlabel': '', 'country':temp['country'],
                                     'md5':temp['tags'], 'sha256': temp['sub-Net'], 'related': tempTI})
                else: 
                    IoCsRP.append({'IoC':i, 'type':'IP', 'score':'Not Found',
                                     'IoCName':'', 'threatlabel': '', 'country':'',
                                     'md5':'', 'sha256':'', 'related': []})
        writeExcellFile(IoCsRP, 'ti')
    return 

def writeExcellFile(IoCs, type):
    global countRow
    global userFileName
    if type == 'r':
        try:
            # if userFileName != 'None':
                # workbook = xlsxwriter.Workbook('{}_Results{}_Reputation.xlsx'.format(userFileName,random.randint(0,10)))
                
            # else : 
            workbook = xlsxwriter.Workbook('Results{}_Reputation.xlsx'.format(random.randint(0,10)))
            userFileName = 'None'
            
            worksheet= workbook.add_worksheet('SHEHANH tool')
            worksheet.write('A1','Type')
            worksheet.write('B1','user_Input')
            worksheet.write('C1','Threat label')
            worksheet.write('D1','Reputation Score')
            worksheet.write('E1','MD5')
            worksheet.write('F1','Sha256')
            
            
            
            for i in IoCs:

                if i['type'] == 'hashe':
                    worksheet.write('A'+str(countRow),'File')
                elif i['type'] == 'domain':
                    worksheet.write('A'+str(countRow),'Domain')
                elif i['type'] == 'IP':
                    worksheet.write('A'+str(countRow),'IP')
                
                worksheet.write('B'+str(countRow),str(i['IoC']))
                worksheet.write('C'+str(countRow),str(i['threatlabel']))
                worksheet.write('D'+str(countRow),str(i['score']))                    
                worksheet.write('E'+str(countRow),str(i['md5']))
                worksheet.write('F'+str(countRow),str(i['sha256']))
                countRow+=1

        finally:
            workbook.close ()
            countRow = 2

    elif type == 'th':
        try:
            # if userFileName != 'None':
            #     workbook = xlsxwriter.Workbook('{}_Results{}_Threat_Huntting.xlsx'.format(userFileName,random.randint(0,10)))
            # else : 
            workbook = xlsxwriter.Workbook('Results{}_Threat_Huntting.xlsx'.format(random.randint(0,99)))
            userFileName = 'None'

            worksheet= workbook.add_worksheet('SHEHANH tool')
            worksheet.write('A1','User Input')
            worksheet.write('B1','Type')
            worksheet.write('C1','Reputation Score')
            worksheet.write('D1','Threat label')
            worksheet.write('E1','Value 1 (MD5/ tags/ Domain)')
            worksheet.write('F1','Value 2 (Sha256/ IP)')
            worksheet.write('G1','Type of Extension')
            worksheet.write('H1','ImportList (import libraray , import funcation)')
            worksheet.write('I1','Size')
            worksheet.write('J1','IoC Name')
            for i in IoCs:
                
                if i['type'] == 'File': 
                    # if checker availbele 
# temp = {
#                                                     'found': 'True', 'type': 'file', 'indicatorScor': 'Malicious/Suspicious',
#                                                     'threatlabel': None, 'indicatorName': 'N/A',
#                                                     'typeTag': None, 'sha256': None,
#                                    }

                    
                    worksheet.write('A'+str(countRow),str(i['IoC']))
                    worksheet.write('B'+str(countRow),str(i['type']))
                    worksheet.write('C'+str(countRow),str(i['score']))
                    worksheet.write('D'+str(countRow),str(i['threatlabel']))
                    worksheet.write('E'+str(countRow),str(i['md5']))
                    worksheet.write('F'+str(countRow),str(i['sha256']))
                    worksheet.write('G'+str(countRow),str(i['typeExtension']))
                    worksheet.write('H'+str(countRow),str(i['importList']))
                    worksheet.write('I'+str(countRow),str(i['size']))
                    worksheet.write('J'+str(countRow),str(i['IoCName']))
                    countRow+=1
                elif i['type'] == 'domain':
                    # IoCsRP.append({'IoC':i, 'type':'domain', 'score':temp['indicatorScor'],
                    #                  'IoCName':'N/A', 'threatlabel': temp['categories'],
                    #                  'md5':temp['tags'], 'sha256':temp['record_A'], 'related': tempTI})
                    worksheet.write('A'+str(countRow),str(i['IoC']))
                    worksheet.write('B'+str(countRow),str(i['type']))
                    worksheet.write('C'+str(countRow),str(i['score']))
                    worksheet.write('D'+str(countRow),str(i['threatlabel']))
                    worksheet.write('E'+str(countRow),str(i['md5']))
                    worksheet.write('F'+str(countRow),str(i['sha256']))
                    countRow+=1
                if i['type'] == 'IP': 
                    # IoCsRP.append({'IoC':i, 'type':'IP', 'score':temp['indicatorScor'],
                    #                  'IoCName':'N/A', 'threatlabel': 'N/A',
                    #                  'md5':temp['tags'], 'sha256': temp['sub-Net'], 'related': tempTI})
                    worksheet.write('A'+str(countRow),str(i['IoC']))
                    worksheet.write('B'+str(countRow),str(i['type']))
                    worksheet.write('C'+str(countRow),str(i['score']))
                    worksheet.write('E'+str(countRow),str(i['md5']))
                    worksheet.write('F'+str(countRow),str(i['sha256']))
                    countRow+=1
                # if the user Input (IoC) has a related value
                # then pars them

                for n in i['related']: 

                    if n['type'] == 'file' and n['found'] == 'True':

                        if 'tags' in n : 
                            worksheet.write('B'+str(countRow),str(n['type']))
                            worksheet.write('C'+str(countRow),str(n['indicatorScor']))
                            worksheet.write('D'+str(countRow),str(n['threatlabel']))
                            worksheet.write('E'+str(countRow),str(n['md5'])) 
                            worksheet.write('F'+str(countRow),str(n['sha256']))
                            worksheet.write('G'+str(countRow),str(n['typeExtension']))
                            worksheet.write('H'+str(countRow),str(i['importList']))
                            worksheet.write('I'+str(countRow),str(n['size']))
                            worksheet.write('J'+str(countRow),str(n['indicatorName']))
                            countRow+=1

                        else: 
                            worksheet.write('B'+str(countRow),str(n['type']))
                            worksheet.write('C'+str(countRow),str(n['indicatorScor']))
                            worksheet.write('D'+str(countRow),str(n['threatlabel']))
                            worksheet.write('E'+str(countRow),str(n['typeTag']))
                            worksheet.write('F'+str(countRow),str(n['sha256']))
                            countRow+=1

                    elif n['type'] == 'domain':

                        if 'categories' in n:
                            worksheet.write('B'+str(countRow),str(n['type']))
                            worksheet.write('C'+str(countRow),str(n['indicatorScor']))
                            worksheet.write('D'+str(countRow),str(n['categories']))
                            worksheet.write('E'+str(countRow),str(n['domain']))
                            worksheet.write('F'+str(countRow),str(n['record_A']))
                            countRow+=1
                        else: 
                            worksheet.write('B'+str(countRow),str(n['type']))
                            worksheet.write('C'+str(countRow),str(n['indicatorScor']))
                            worksheet.write('E'+str(countRow),str(n['IoC']))
                            worksheet.write('F'+str(countRow),str(n['typeTag']))
                            countRow+=1

                    elif n['type']== 'IP':

                        if 'tags' in  n: 

                            worksheet.write('B'+str(countRow),'ip')
                            worksheet.write('C'+str(countRow),str(n['indicatorScor']))   
                            worksheet.write('D'+str(countRow),str(n['tags']))
                            worksheet.write('E'+str(countRow),str(n['IoC']))
                            worksheet.write('F'+str(countRow),str(['sub-Net']))
                            countRow+=1
                        else: 

                            worksheet.write('B'+str(countRow),'ip')
                            worksheet.write('C'+str(countRow),str(n['indicatorScor']))   
                            worksheet.write('D'+str(countRow),str(n['typeTag']))
                            worksheet.write('E'+str(countRow),str(n['IoC']))
                            countRow+=1
                            
                        
        except :
            print(Exception)
        finally:
            workbook.close ()
            countRow = 2
    elif type == 'ti':
        try:
            # if userFileName != 'None':
            #     workbook = xlsxwriter.Workbook('{}_Results{}_Threat_Intelligence.xlsx'.format(userFileName,random.randint(0,10)))
            # else : 
            workbook = xlsxwriter.Workbook('Results{}_Threat_Intelligence.xlsx'.format(random.randint(0,99)))
            userFileName = 'None'
            
            workbook = xlsxwriter.Workbook('Results{}_Threat_Intelligence.xlsx'.format(random.randint(0,99)))
            worksheet= workbook.add_worksheet('SHEHANH tool')
            worksheet.write('A1','User Input')
            worksheet.write('B1','Type')
            worksheet.write('C1','Reputation Score')
            worksheet.write('D1','IoC Name')
            worksheet.write('E1','Threat label')
            worksheet.write('F1','Type extension')
            worksheet.write('G1','Value 1 (MD5/ Domain/ IP)')
            worksheet.write('H1','Value 2 (Sha1/ IP/ Sub-net)')
            worksheet.write('I1','Value 3 (Sha256/ Creation date/ Country)')
            worksheet.write('J1','ImportList (import libraray , import funcation)')
            worksheet.write('K1','Size')
            worksheet.write('L1','Section')
            worksheet.write('M1','Tags')

            for i in IoCs:
                
                worksheet.write('A'+str(countRow),i['IoC'])
                worksheet.write('B'+str(countRow),i['type'])
                worksheet.write('C'+str(countRow),i['score'])

                if i['type'] == 'File':
                    worksheet.write('D'+str(countRow),str(i['IoCName']))
                    worksheet.write('E'+str(countRow),str(i['threatlabel']))
                    worksheet.write('F'+str(countRow),str(i['typeExtension']))
                    worksheet.write('G'+str(countRow),str(i['md5']))
                    worksheet.write('H'+str(countRow),str(i['sha1']))
                    worksheet.write('I'+str(countRow),str(i['sha256']))
                    worksheet.write('J'+str(countRow),str(i['importList']))
                    worksheet.write('K'+str(countRow),str(i['size']))
                    worksheet.write('L'+str(countRow),str(i['sections']))
                    worksheet.write('M'+str(countRow),str(i['tags']))
                elif i['type'] == 'IP':
                    worksheet.write('H'+str(countRow),str(i['sha256'])) #sub-net
                    worksheet.write('I'+str(countRow),str(i['country']))
                    worksheet.write('M'+str(countRow),str(i['md5'])) #tags

                elif i['type'] == 'domain': 

                    worksheet.write('E'+str(countRow),str(i['threatlabel']))
                    worksheet.write('G'+str(countRow),str(i['sha256'])) #IP
                    worksheet.write('I'+str(countRow),str(i['creation_date']))
                    worksheet.write('M'+str(countRow),str(i['tags']))
                countRow+=1

                for n in i['related']: 

                    worksheet.write('B'+str(countRow),n['type'])
                    worksheet.write('C'+str(countRow),n['indicatorScor'])

                    if n['type'] == 'file':
                        worksheet.write('D'+str(countRow),str(n['indicatorName']))
                        worksheet.write('E'+str(countRow),str(n['threatlabel']))
                        worksheet.write('F'+str(countRow),str(n['typeExtension']))
                        worksheet.write('G'+str(countRow),str(n['md5']))
                        worksheet.write('H'+str(countRow),str(n['sha1']))
                        worksheet.write('I'+str(countRow),str(n['sha256']))
                        worksheet.write('J'+str(countRow),str(n['importList']))
                        worksheet.write('K'+str(countRow),str(n['size']))
                        worksheet.write('L'+str(countRow),str(n['sections']))
                        worksheet.write('M'+str(countRow),str(n['tags']))
                    elif n['type'] == 'domain':
  
                        worksheet.write('E'+str(countRow),str(n['categories']))
                        worksheet.write('G'+str(countRow),str(n['domain']))
                        worksheet.write('H'+str(countRow),str(n['record_A']))
                        worksheet.write('I'+str(countRow),str(n['creation_date']))
                        worksheet.write('M'+str(countRow),str(n['tags']))

                    elif n['type'] == 'ip':
                        worksheet.write('G'+str(countRow),str(n['IoC']))
                        worksheet.write('H'+str(countRow),str(n['sub-Net']))
                        worksheet.write('I'+str(countRow),str(n['country']))
                        worksheet.write('M'+str(countRow),str(n['tags']))
                    countRow+=1


        except :
            print(Exception)
        finally:
            workbook.close ()
            countRow = 2

def displayLive(IoCs, type): 
    # print(IoCs)
    if type == 'th': 
            # a dict inside array with keys ('Input','Type','Reputation_Score','Threat_label', 'Value 2 (Sha256/IP/domain)')
        
        IoCSort = []

        for i in IoCs:
                
                IoCSubSort={'Input': i['IoC'], 'type':i['type'],'Reputation_Score':i['score'],
                            'Value (Sha256/IP/domain)':i['sha256']
                }
                # ,
                #              'Threat_label':i['threatlabel']
                IoCSort.append(IoCSubSort)


                for n in i['related']: 
                    
                    IoCSubSort2=None
                    if n['type'] == 'file':
                        IoCSubSort2 = { 'Input':'', 'type' : n['type'], 
                                    'Reputation_Score':n['indicatorScor'], 
                                    'Value (Sha256/IP/domain)':n['sha256']
                        }
                    elif n['type'] == 'domain':
                        IoCSubSort2 = { 'Input':'', 'type' : n['type'], 
                                    'Reputation_Score':n['indicatorScor'], 
                                    'Value (Sha256/IP/domain)':'' 
                        }

                        if 'IoC' in n :
                            IoCSubSort2['Value (Sha256/IP/domain)'] = n['IoC']
                        else: 
                            IoCSubSort2['Value (Sha256/IP/domain)'] = n['domain']


                    elif n['type']== 'IP':
                        IoCSubSort2 = { 'Input':'', 'type' :'ip', 
                                    'Reputation_Score':n['indicatorScor'], 
                                    'Value (Sha256/IP/domain)':'' 
                        }
# 'Threat_label':'N/A',
                        if 'IoC' in n :
                            IoCSubSort2['Value (Sha256/IP/domain)'] = n['IoC']
                        else:
                            IoCSubSort2['Value (Sha256/IP/domain)'] = 'error 34'
                    if IoCSubSort2 == None : 
                        continue
                    IoCSort.append(IoCSubSort2)
        
        
        header = IoCSort[0].keys()
        rows =  [x.values() for x in IoCSort]
        print(tabulate(rows, header))
    elif type == 'r':
        IoCSort = []

        for i in IoCs:

                IoCSubSort={'Input': str(i['input']), 'type':str(i['type']),'Reputation_Score':str(i['scor']),
                            'Value (Sha256/ IP/ Sub-Net)':str(i['sha256'])}
                IoCSort.append(IoCSubSort)
        
        
        header = IoCSort[0].keys()
        rows =  [x.values() for x in IoCSort]
        print(tabulate(rows, header))

    pass

def main():
    userrun = True
    printLogo()
    print("Let's Hunt !\n\n")
    type = None
    chosemodel= False 
    Live = True
    searchM= False 
    usermodule = 'Write your order > '
    # SInput =None
    if(not readAPIKeyFile()):
            Input= input ( "You dont have API key, write your API keys " )
            keysList = Input.split()
            # keysList = Input.split(" ",1)[1]
            # keysList = str(keysList).split(" ")
            
            setAPIKeyFile(keysList)
            try: 
                if readAPIKeyFile():
                    print("Your keys have been saved ")
                    
                else:
                    print('Error during set your API Keys ')
            except: 
                print('Error during set your API Keys ')
                # print(ApiKeylist)
    
    userIoCs = {'hashes':[], 'domains':[], 'Ips':[]}
    while (userrun):
        if not chosemodel:
            print('Choose your search Module\n')
            print( "Availbale modules\n" )

            M = [['r ', 'Retrieve reputation  of IoCs'],
                    ['th','Retrieve all IoCs realted to your IoC, with Maximem 30 IoCs realted to one IoC, '],
                    ['','WITHOUT full details for the related IoCs'],
                    ['',''],
                    ['ti','Retrieve all IoCs realted to yur IoC, with Maximem 30 IoCs realted to one IoC, '],
                    ['','WITH full detailsfor the related IoCs. The results will be exported as a file']
                ]
            print (tabulate(M, headers=['Module', 'Description']))
            print('\n \n ')
                # print (tabulate(TOS, headers=['Type Of Search', 'Description']))
            print('Your order should be wretin as following to use Module ')
            print('use r \n')
        elif ( not searchM):
                print('\nChoose your search type\n')
                print( "search types\n" )

                IT = [['d', 'The [IoC] will be a domains'],
                    ['hash',' The [IoC] will be a hashs'],
                    ['ip','The [IoC] will be an IPs'],
                    ['file','The [IoC] will be a file path, The file extension should be one of (xlsx, csv), the colums should be have'],
                    ['','one of (Hashes, Ips, Domains) on the first row of the file and conntain all IoCs you wnat to search for them'],
                    ['export','To export the results as excell fille <more details> write a [ set export true ]. However, the default is not to export them '],
                    ['key','To set new API keys']
                    ]
                print (tabulate(IT, headers=['Set type', 'Description']))
                print('\n \n ')
            
                # print (tabulate(TOS, headers=['Type Of Search', 'Description']))
                print('your order should be wetin as following to search type\n')
                print('set ip 1.1.1.1')
                print('set d example.com\n')

        userInputs = input ("\n{}".format(usermodule))

        userInput = str(userInputs)
        userInput = userInput.split(" ")
        # print(userInput)
        if (len(userInput) < 2 and (userInput[0] =='h' or userInput[0] =='help' or userInput[0] =='?') and  userInput[0] !='run'): 
            # if (userInput[0] == ('menu' or 'Menu' or 'm' or 'M' )): 
                print('\n \n IN Shihanah tool, you have to choose you Module first, then your set valuse,then write run')
                print( "\nAvailbale modules\n" )
                M = [['r ', 'Retrieve reputation of IoC'],
                        ['th','Retrieve all IoCs realted to yur IoC, with Maximem 30 IoCs realted to one IoC, '],
                        ['','WITHOUT full details for the related IoCs'],
                        ['',''],
                        ['ti','Retrieve all IoCs realted to yur IoC, with Maximem 30 IoCs realted to one IoC, '],
                        ['','WITH full detailsfor the related IoCs']
                    ]
                print (tabulate(M, headers=['Module', 'Description']))
                print('\n \n ')
                print('Your order should be wretin as following to use Module ')
                print('use r \n')
                print('use th \n')

                print( "Availbale types\n" )

                IT = [['d', 'The [IoC] will be a domains'],
                    ['hash',' The [IoC] will be a hashs'],
                    ['ip','The [IoC] will be an IPs'],
                    ['file','The [IoC] will be a file path, The file extension should be one of (xlsx, csv), the colums should be have'],
                    ['','one of (Hashes, Ips, Domains) on the first row of the file and conntain all IoCs you wnat to search for them'],
                    ['export','To export the results as excell fille <more details> write a [ set export true ]. However, the default is not to export them '],
                    ['key','To set new API keys']
                    ]
                print (tabulate(IT, headers=['Set type', 'Description']))
                print('\n \n ')
                print('your order should be wretin as following to set values or search type\n')
                print('set ip 1.1.1.1')
                print('set d example.com')
                print('set export true\n')
            # print("wrong input")
            # continue
        elif (userInput[0] == 'use'): 
            # userInput = userInput.split("use",1)[1]
            if (userInput[1] == 'r' or userInput[1] == 'th' or userInput[1] == 'ti'): 
                if userInput[1] == 'r' :
                    type = 'r'
                    usermodule = 'Reputation Module > '
                    print('')
                elif userInput[1] == 'th':
                    type = 'th'
                    usermodule = 'Threat Hunting Module > '
                    # print('Your Module is ')
                else : 
                    type = 'ti'
                    usermodule = 'Threat Intelligence Module > '
                    # print('Your Module is Threat ')
                searchM = False  
                chosemodel = True
                
                continue
            else: 
                print('Wrong module, write use then Module name ,\n')
                chosemodel = False
                continue
        elif (userInput[0] =='set' ):
            if chosemodel :
                # userInput = userInput.split("set",1)[1]
                # userInput2= userInput.split(" ")
                if ( userInput[1] == 'd' or userInput[1] == 'hash' or userInput[1] == 'file' or userInput[1] == 'key' or  userInput[1] == 'export'or userInput[1] == 'ip') :
                        if (len(userInput) <= 2): 
                            print('write your valuse')
                            continue
                        order = userInput[1]
                        if userInput[1] == 'd':
                            # domainsList = userInput.split("d",1)[1]  
                            # domainsList = str(domainsList)
                            # domainsList = domainsList.split(" ")
                            del userInput[0:2]
                            for i in userInput: 
                                if len(i)>=3:
                                    userIoCs['domains'].append(i)
                            
                        elif userInput[1] == 'ip':
                            # ipslist = userInput.split("ip",1)[1]
                            # ipslist = str(ipslist)
                            # ipslist = ipslist.split(" ")
                            del userInput[0:2]
                            for i in userInput: 
                            # for i in ipslist: 
                                if len(i)>=6:
                                    userIoCs['Ips'].append(i)
                            
                        elif userInput[1] == 'hash':
                            del userInput[0:2]
                            for i in userInput: 
                                if len(i)>=6:
                                    userIoCs['hashes'].append(i)
                            # print('The value has been saved')
                        elif userInput[1] =='file':
                            del userInput[0:2]
                            filepath  = 'first time'
                            counts=1
                            if (len(userInput)>=2):
                                for i in userInput:
                                    if counts ==1:
                                        filepath = i + ' '
                                        counts +=1
                                        continue
                                    if ( counts == len(userInput)):
                                        filepath += i
                                        break
                                    filepath = filepath +i 
                                    filepath = filepath + " "
                                    counts +=1
                            
                            filepath= str(filepath)
                            filepath = filepath.replace('\'', '')
                            filepath = filepath.replace('\"', '')
                            try:
                                userIoCs = readFile(filepath)
                                pruserIoCs = str(userIoCs)
                                print('user IoCs'+ pruserIoCs)
                                if userIoCs['error'] == 1 or len(userIoCs) <= 1 :
                                    print ('Kindly, Check the file type or inputs\n')
                                    print ('Enter --help, to get the mune\n')
                                # else:
                                    
                                    # print('The value has been saved')
                            except Exception as a:
                                print(a)
                        elif userInput[1] =='export' :
                            if userInput[2] == 'true':
                                Live= False 
                                print('The results will be exported as file')
                            else:
                                print('The results will not be exported as file')
                                Live= True
                        elif userInput[1] =='key' : 
                            del userInput[0:2]
                            setAPIKeyFile(userInput)
                            try: 
                                if readAPIKeyFile():
                                    print("\nYour keys have been saved")
                                    
                                else:
                                    print('\nError during set your API Keys')
                            except: 
                                print('\nError during set your API Keys')
                        if (order !='key') and (order != 'export'): 
                            print('Your set has been save, please write run / set for further orders\n')
                        searchM = True
                        continue
                else : 
                    print('unknown set, write a correct set option \n')
                    continue
            else: 
                print('You dont chose a module, chose your module \n')
                continue
        elif(userInput[0] =='run' ):
            if(chosemodel and searchM):
                try:
                    print('\n')
                    check(type, userIoCs , Live)
                    
                    userIoCs['hashes'].clear()
                    userIoCs['domains'].clear()
                    userIoCs['Ips'].clear()
                    # Live = True
                    searchM = True 
                    print('Validate the results as they may be false positives from Virus Total\n')
                    if ( not Live):
                        print('Your results has been saved, The file saved in program directory\n')
                    print('set new values\n')
                    continue
                except Exception as e: 
                    print('Check your connection or the bellow error ')
                    print(e)
                    continue
            else: 
                if not chosemodel: 
                    print('You dodn\'t chose your module, write use then you module\n')
                else: 
                    print('You dont have a set value, write your set order with value \n ')
        elif (userInput[0] =='exit' ): 
            print('Have nice hunting, Goodbye \n' )
            userrun = False
        
       
               
def printLogo(): 

    logo= ''''
⠀⠀⠀⠀⠀⠀⣠⣴⣾⣿⣿⣿⣿⣿⣿⣶⣤⣄⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⣠⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣄⠀⠀⠀⠀⠀
⠀⠀⣠⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣆⠉⠉⢉⣿⣿⣿⣷⣦⣄⡀⠀
⠀⠚⢛⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡄
⠀⢠⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠿⠿⠿⠿⣿⡇
⢀⣿⡿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠋⠁⠀⠀⠀⠀⠀⠀⠈⠃
⠸⠁⢀⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠏⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⢸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡏⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⣿⣿⣿⡿⣿⣿⣿⣿⣿⣿⠁⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠹⣿⣿⡇⠈⠻⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠈⠻⡇⠀⠀⠈⠙⠿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
     _______. __    __   _______  __    __       ___      .__   __.      ___       __    __  
    /       ||  |  |  | |   ____||  |  |  |     /   \     |  \ |  |     /   \     |  |  |  | 
   |   (----`|  |__|  | |  |__   |  |__|  |    /  ^  \    |   \|  |    /  ^  \    |  |__|  | 
    \   \    |   __   | |   __|  |   __   |   /  /_\  \   |  . `  |   /  /_\  \   |   __   | 
.----)   |   |  |  |  | |  |____ |  |  |  |  /  _____  \  |  |\   |  /  _____  \  |  |  |  | 
|_______/    |__|  |__| |_______||__|  |__| /__/     \__\ |__| \__| /__/     \__\ |__|  |__| 
                                                                                             
'''

    Logo2=''''
    

     _______. __    __   __   __    __       ___      .__   __.      ___       __    __  
    /       ||  |  |  | |  | |  |  |  |     /   \     |  \ |  |     /   \     |  |  |  | 
   |   (----`|  |__|  | |  | |  |__|  |    /  ^  \    |   \|  |    /  ^  \    |  |__|  | 
    \   \    |   __   | |  | |   __   |   /  /_\  \   |  . `  |   /  /_\  \   |   __   | 
.----)   |   |  |  |  | |  | |  |  |  |  /  _____  \  |  |\   |  /  _____  \  |  |  |  | 
|_______/    |__|  |__| |__| |__|  |__| /__/     \__\ |__| \__| /__/     \__\ |__|  |__| 
                                                                                         

                                                                                             
    '''
    
    print(Logo2)



if __name__ == "__main__":
     main()

        